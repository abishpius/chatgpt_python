# Import packages
import pandas as pd
import numpy as np
import openai
import PyPDF2
import openpyxl
import docx
# import tiktoken
import os
import streamlit as st
from openai.embeddings_utils import distances_from_embeddings
import base64


## Params ##
# Set Key
openai.api_key = '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'

# Output Limit
max_tokens = 2000

# Text Reduction
split_tokens = 500

# Other Openai args
temp = 0.5
stop_phrase = ["Hope this helps."]

## END Params ##


## HELPER FUNCTIONS ##

def preprocess_text(serie):
    serie = serie.replace('\tNone', '')
    serie = serie.replace('None\t', '')
    serie = serie.replace('\n', ' ')
    serie = serie.replace('\\n', ' ')
    serie = serie.replace('  ', ' ')
    serie = serie.replace('  ', ' ')
    return serie

def pdf_to_text(pdf_file):
  # Open the PDF file directly
    pdf_reader = PyPDF2.PdfReader(pdf_file)
    text = ''
    for page_num in range(len(pdf_reader.pages)):
        page = pdf_reader.pages[page_num]
        text += page.extract_text()
        
    return text

def excel_to_text(excel_file):
  # Open the Excel file
    wb = openpyxl.load_workbook(excel_file)
  # Create an empty string to store the converted text
    converted_text = ""
  # Iterate over each sheet in the Excel file
    for sheet_name in wb.sheetnames:
        # Get the sheet object
        sheet = wb[sheet_name]
      # Add the sheet name to the converted text string
        converted_text += f"{sheet_name}\n"
      # Iterate over each row in the sheet
        for row in sheet.iter_rows(values_only=True):
          # Concatenate the row to the converted text string, separated by tabs
            row_str = "\t".join(str(cell) for cell in row) + "\n"
            converted_text += row_str
      # Add a newline separator between sheets
        converted_text += "\n"
  # Return the converted text string
    return converted_text

def read_docx_file(file_path):
    doc = docx.Document(file_path)
    full_text = []
    for para in doc.paragraphs:
        full_text.append(para.text)
    return '\n'.join(full_text)

def read_text_file(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        text = f.read()
    return text

# Function to split the text into chunks of a maximum number of tokens
def split_into_many(text, max_tokens = split_tokens):
    # Split the text into sentences
    sentences = text.split('. ')
    # Get the number of tokens for each sentence
#     n_tokens = [len(tokenizer.encode(" " + sentence)) for sentence in sentences]
    n_tokens = [len(" " + sentence)/4 for sentence in sentences]
    chunks = []
    tokens_so_far = 0
    chunk = []
    # Loop through the sentences and tokens joined together in a tuple
    for sentence, token in zip(sentences, n_tokens):
        # If the number of tokens so far plus the number of tokens in the current sentence is greater 
        # than the max number of tokens, then add the chunk to the list of chunks and reset
        # the chunk and tokens so far
        if tokens_so_far + token > max_tokens:
            chunks.append(". ".join(chunk) + ".")
            chunk = []
            tokens_so_far = 0
        # If the number of tokens in the current sentence is greater than the max number of 
        # tokens, go to the next sentence
        if token > max_tokens:
            continue
        # Otherwise, add the sentence to the chunk and add the number of tokens to the total
        chunk.append(sentence)
        tokens_so_far += token + 1

    return chunks

def create_context(
    question, df, max_len=1800, size="ada"
):
    """
    Create a context for a question by finding the most similar context from the dataframe
    """

    # Get the embeddings for the question
    q_embeddings = openai.Embedding.create(input=question, engine='text-embedding-ada-002')['data'][0]['embedding']

    # Get the distances from the embeddings
    df['distances'] = distances_from_embeddings(q_embeddings, df['embeddings'].values, distance_metric='cosine')

    returns = []
    cur_len = 0

    # Sort by distance and add the text to the context until the context is too long
    for i, row in df.sort_values('distances', ascending=True).iterrows():
        
        # Add the length of the text to the current length
        cur_len += row['n_tokens'] + 4
        
        # If the context is too long, break
        if cur_len > max_len:
            break
        
        # Else add it to the text that is being returned
        returns.append(row["text"])

    # Return the context
    return "\n\n###\n\n".join(returns)

def answer_question(
    df,
    model="text-davinci-003",
    question="What are the controls related to access?",
    max_len=1800,
    size="ada",
    debug=False,
    max_tokens=max_tokens,
    stop_sequence=stop_phrase
):
    """
    Answer a question based on the most similar context from the dataframe texts
    """
    context = create_context(
        question,
        df,
        max_len=max_len,
        size=size,
    )
    # If debug, print the context model uses
    if debug:
        print("Context:\n" + context)
        print("\n\n")

    try:
        # Create a completions using the question and context
        response = openai.Completion.create(
            prompt=f"Answer the question based on the context below, and supplement it with general information if applicable \n\nContext: {context}\n\n---\n\nQuestion: {question}\nAnswer:",
            temperature=temp,
            max_tokens=max_tokens,
            top_p=1,
            frequency_penalty=0,
            presence_penalty=0,
            stop=stop_sequence,
            model=model,
        )
        return response["choices"][0]["text"].strip()
    except Exception as e:
        print(e)
        return ""

def convert_df(df):
        # IMPORTANT: Cache the conversion to prevent computation on every rerun
    return df.to_csv().encode('utf-8')    
## END HELPER FUNCTIONS ##

## BEGIN STREAMLIT PORTION ##

st.title("Risk Assist AI")

store_queries = []
store_responses = []
flag = False

upload_file = st.file_uploader("Please upload a file here", type=["pdf", "docx", "xlsx", "xlsm"])

if upload_file is not None:
    filetype = upload_file.name.split('.')[-1]
    filename =  upload_file.name.split('.')[0]
    # Addres File Types
# Check pdf
    if filetype =='pdf':
        text = pdf_to_text(upload_file)

# Check Excel
    elif filetype == 'xlsm' or filetype == 'xlsx':
        text = excel_to_text(upload_file)

# Check if DocX
    elif filetype == 'docx':
        text = read_docx_file(upload_file)

# Assume text by default
    else:
        text = read_text_file(upload_file)

# Add Preprocessing
    text = preprocess_text(text)
# Create a dataframe from the list of texts
    df = pd.DataFrame([[filename,text]], columns = ['fname', 'text'])
# Set the text column to be the raw text with the newlines removed
    df['text'] = df.fname + ". " + df.text
    
#     # Load the cl100k_base tokenizer which is designed to work with the ada-002 model
#     tokenizer = tiktoken.get_encoding("cl100k_base")

    # df = pd.read_csv('processed/scraped.csv', index_col=0)
    df.columns = ['title', 'text']

    # Tokenize the text and save the number of tokens to a new column
    df['n_tokens'] = df.text.apply(lambda x: int(len(x)/4)+1)#df['n_tokens'] = df.text.apply(lambda x: len(tokenizer.encode(x)))
    
        
    shortened = []

    # Loop through the dataframe
    for row in df.iterrows():
        # If the text is None, go to the next row
        if row[1]['text'] is None:
            continue
        # If the number of tokens is greater than the max number of tokens, split the text into chunks
        if row[1]['n_tokens'] > max_tokens:
            shortened += split_into_many(row[1]['text'])
        # Otherwise, add the text to the list of shortened texts
        else:
            shortened.append( row[1]['text'] )
    
    df = pd.DataFrame(shortened, columns = ['text'])  
    df['n_tokens'] = df.text.apply(lambda x: int(len(x)/4)+1)#df['n_tokens'] = df.text.apply(lambda x: len(tokenizer.encode(x)))
    
    st.markdown('#### Sample of Uploaded Text Object Being Parsed for Fine Tuning')
#     st.divider()
    st.dataframe(df.head(3))
    
    # Train ADA Model
    if st.button("Begin Document Fine Tuning, ONLY CLICK IF SURE CORRECT FILE IS UPLOADED!"):
        df['embeddings'] = df.text.apply(lambda x: openai.Embedding.create(input=x, engine='text-embedding-ada-002')['data'][0]['embedding'])
        df.to_csv('local.csv')
        
        st.markdown('###### Fine Tuning Was Successful!')
        flag = True
        with open('output.txt', 'w') as file:
            file.write('')

    
    st.markdown('#### Run Queries on Document')
    st.markdown('*NOTE: Only works after fine tuning step above. May take some time to generate the answer, please do not refresh the page in the meantime.*')
    
    txt_input = st.text_input('Please Type Your Query Here')  
    if txt_input:
        # So as to not reload embeddings everytime
        df = pd.read_csv('local.csv', index_col = 0)
        df['embeddings'] = df['embeddings'].apply(eval).apply(np.array)
        ans = answer_question(df,  question=txt_input)
        store_responses.append(ans)
        store_queries.append(txt_input)
        result = pd.DataFrame({'Query': store_queries,
                                  'Answers': store_responses})
        with open('output.txt', 'a') as f:  # use 'a' mode to append to file
            for x, y in zip(store_queries, store_responses):
                f.write(f"{x}\t{y}\n")
        st.write('Answer:', ans)
        st.markdown('##### Feel free to try another query or download all your queries and answers below.')
        
    
    if st.button("Prepare Q and A for Download"):
        res = pd.read_csv('output.txt', sep = '\t', index_col = 0, encoding='Windows-1252')
        st.markdown('##### Sample of Output:')
        st.dataframe(res.head(3))
        csv = convert_df(res)

        st.download_button(
            label="Download data as CSV",
            data=csv,
            file_name='Q_and_A.csv',
            mime='text/csv',
        )
#         # Open the file in read mode
#         with open('output.txt', 'r') as f:
#             # Read the contents of the file into a string variable
#             file_contents = f.read()
        
#         st.download_button('File Ready, Download Here', file_contents)
                 
    
    
        
        
        